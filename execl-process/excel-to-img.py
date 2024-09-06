from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import colorsys


def excel_to_image(excel_file, sheet_name=None):
    # 加载工作簿
    wb = load_workbook(excel_file)

    # 如果没有指定sheet_name，使用活动工作表
    ws = wb[sheet_name] if sheet_name else wb.active

    # 获取工作表的尺寸
    max_row = ws.max_row
    max_col = ws.max_column

    # 设置单元格大小（像素）
    cell_width = 100
    cell_height = 20

    # 创建一个白色背景的图像
    img_width = max_col * cell_width
    img_height = max_row * cell_height
    img = Image.new("RGB", (img_width, img_height), color="white")

    # 创建绘图对象
    draw = ImageDraw.Draw(img)

    # 使用一个简单的字体
    font = ImageFont.load_default()

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws[f"{get_column_letter(col)}{row}"]

            # 获取单元格位置
            x = (col - 1) * cell_width
            y = (row - 1) * cell_height

            # 绘制单元格背景色
            if cell.fill.start_color.index != "00000000":
                fill_color = cell.fill.start_color.rgb
                if fill_color:
                    r, g, b = (
                        int(fill_color[2:4], 16),
                        int(fill_color[4:6], 16),
                        int(fill_color[6:8], 16),
                    )
                    draw.rectangle(
                        [x, y, x + cell_width, y + cell_height], fill=(r, g, b)
                    )

            # 绘制单元格边框
            draw.rectangle([x, y, x + cell_width, y + cell_height], outline="black")

            # 绘制单元格内容
            if cell.value:
                draw.text((x + 5, y + 5), str(cell.value), fill="black", font=font)

    return img


# 使用示例
excel_file = "output_reversed_daily_month.xlsx"
sheet_name = "Sheet"  # 如果不指定，将使用活动工作表
img = excel_to_image(excel_file, sheet_name)
img.save("output.png")
