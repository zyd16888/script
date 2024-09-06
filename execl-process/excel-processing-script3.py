import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from collections import defaultdict


def get_color(value):
    if 90 <= value <= 100:
        return "8B4513"  # 棕色
    elif 80 <= value < 90:
        return "FF0000"  # 红色
    elif 70 <= value < 80:
        return "FFA500"  # 橙色
    elif 60 <= value < 70:
        return "FFFF00"  # 黄色
    elif 50 <= value < 60:
        return "008000"  # 绿色
    elif 40 <= value < 50:
        return "0000FF"  # 蓝色
    else:
        return None


def process_excel(input_file, output_file):
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active

    # Create a new workbook for output
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active

    # Collect and organize data
    data = defaultdict(lambda: defaultdict(dict))
    months = set()
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        date_str, *values = row
        if isinstance(date_str, datetime):
            date = date_str
        else:
            try:
                date = datetime.strptime(date_str, "%m/%d")
            except ValueError:
                print(f"Warning: Could not parse date {date_str}. Skipping.")
                continue

        month = date.strftime("%B")
        months.add(month)

        for i, value in enumerate(values):
            time = timedelta(minutes=5 * i)
            hour = (date + time).strftime("%H:00")
            data[month][hour][date.day] = value

    # Sort months
    months = sorted(months, key=lambda m: datetime.strptime(m, "%B"))

    # Write headers
    ws_output.cell(row=1, column=1, value="Hour")
    ws_output.cell(row=1, column=2, value="Day")
    for col, month in enumerate(months, start=3):
        ws_output.cell(row=1, column=col, value=month)

    # Write data
    current_row = 2
    for hour in sorted(set.union(*[set(data[m].keys()) for m in months])):
        start_row = current_row
        for minute in range(0, 60, 5):
            time = f"{hour[:2]}:{minute:02d}"
            ws_output.cell(row=current_row, column=2, value=time)

            for col, month in enumerate(months, start=3):
                for day, value in data[month][hour].items():
                    cell = ws_output.cell(row=current_row, column=col)
                    cell.value = value
                    color = get_color(value)
                    if color:
                        cell.fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )

            current_row += 1

        # Merge hour cells
        ws_output.merge_cells(
            start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1
        )
        ws_output.cell(row=start_row, column=1, value=hour)

    wb_output.save(output_file)


# Usage
input_file = r"新街口站255负荷.xlsx"
output_file = "output_reversed.xlsx"
process_excel(input_file, output_file)
