import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta


def get_color(value):
    if 90 <= value <= 100:
        return "8B4513"  # 棕色
    elif 80 <= value < 90:
        return "FF0000"  # 红色
    elif 70 <= value < 80:
        return "FFA500"  # 橙色
    elif 60 <= value < 70:
        return "FFFF00"  # 黄色
    elif 50 <= value < 60:
        return "008000"  # 绿色
    elif 40 <= value < 50:
        return "0000FF"  # 蓝色
    else:
        return None


def process_excel(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Insert a new column at the beginning
    ws.insert_cols(1)

    # Add month information
    current_month = None
    start_row = None

    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=2)

        if isinstance(date_cell.value, datetime):
            date = date_cell.value
        elif isinstance(date_cell.value, str):
            try:
                date = datetime.strptime(date_cell.value, "%m/%d")
            except ValueError:
                print(f"Warning: Could not parse date in row {row}. Skipping.")
                continue
        else:
            print(f"Warning: Unexpected date format in row {row}. Skipping.")
            continue

        if current_month != date.month:
            if current_month is not None:
                ws.merge_cells(
                    start_row=start_row, start_column=1, end_row=row - 1, end_column=1
                )
            current_month = date.month
            start_row = row
            ws.cell(row=row, column=1, value=date.strftime("%B"))

    # Merge the last month's cells
    if start_row is not None:
        ws.merge_cells(
            start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1
        )

    # Color cells based on values
    for row in ws.iter_rows(
        min_row=2, min_col=3, max_row=ws.max_row, max_col=ws.max_column
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                color = get_color(cell.value)
                if color:
                    cell.fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

    wb.save(output_file)


# Usage
input_file = r"新街口站255负荷.xlsx"
output_file = "output.xlsx"
process_excel(input_file, output_file)
