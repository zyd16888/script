import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from collections import defaultdict
from calendar import month_abbr


# 数值上色，转化行列


def get_color(value):
    if 90 <= value <= 100:
        return "8B4513"  # 棕色
    elif 80 <= value < 90:
        return "FF0000"  # 红色
    elif 70 <= value < 80:
        return "FFA500"  # 橙色
    elif 60 <= value < 70:
        return "FFFF00"  # 黄色
    elif 50 <= value < 60:
        return "008000"  # 绿色
    elif 40 <= value < 50:
        return "0000FF"  # 蓝色
    else:
        return None


def process_excel(input_file, output_file):
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active

    # Create a new workbook for output
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active

    # Collect and organize data
    data = defaultdict(lambda: defaultdict(dict))
    dates = []
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        date_str, *values = row
        if isinstance(date_str, datetime):
            date = date_str
        else:
            try:
                date = datetime.strptime(date_str, "%m/%d")
            except ValueError:
                print(f"Warning: Could not parse date {date_str}. Skipping.")
                continue

        dates.append(date)

        for i, value in enumerate(values):
            time = timedelta(minutes=5 * i)
            hour = (date + time).strftime("%H:00")
            minute = (date + time).strftime("%H:%M")
            data[hour][minute][date] = value

    # Sort dates
    dates.sort()

    # Write month headers
    ws_output.insert_rows(1)
    current_month = None
    start_col = 3
    for col, date in enumerate(dates, start=3):
        if date.month != current_month:
            if current_month is not None:
                ws_output.merge_cells(
                    start_row=1, start_column=start_col, end_row=1, end_column=col - 1
                )
                ws_output.cell(row=1, column=start_col, value=month_abbr[current_month])
            current_month = date.month
            start_col = col

    # Merge the last month
    ws_output.merge_cells(
        start_row=1, start_column=start_col, end_row=1, end_column=col
    )
    ws_output.cell(row=1, column=start_col, value=month_abbr[current_month])

    # Write date headers
    for col, date in enumerate(dates, start=3):
        ws_output.cell(row=2, column=col, value=date.strftime("%m/%d"))

    # Write "Hour" and "Minute" headers
    ws_output.cell(row=2, column=1, value="Hour")
    ws_output.cell(row=2, column=2, value="Minute")

    # Write data (adjust row numbers)
    current_row = 3
    for hour in sorted(data.keys()):
        start_row = current_row
        for minute in sorted(data[hour].keys()):
            ws_output.cell(row=current_row, column=1, value=hour)
            ws_output.cell(row=current_row, column=2, value=minute)

            for col, date in enumerate(dates, start=3):
                value = data[hour][minute].get(date)
                if value is not None:
                    cell = ws_output.cell(row=current_row, column=col)
                    cell.value = value
                    color = get_color(value)
                    if color:
                        cell.fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )

            current_row += 1

        # Merge hour cells
        ws_output.merge_cells(
            start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1
        )

    wb_output.save(output_file)


# Usage
input_file = r"新街口站255负荷.xlsx"
output_file = "output_reversed_daily_month.xlsx"
process_excel(input_file, output_file)
