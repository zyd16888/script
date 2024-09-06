from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import colorsys


def excel_to_image(excel_file, sheet_name=None):
    # 加载工作簿
    wb = load_workbook(excel_file)

    # 如果没有指定sheet_name，使用活动工作表
    ws = wb[sheet_name] if sheet_name else wb.active

    # 获取工作表的尺寸
    max_row = ws.max_row
    max_col = ws.max_column

    # 设置单元格大小（像素）
    cell_width = 100
    cell_height = 30

    # 设置坐标宽度
    coord_width = 40

    # 创建一个白色背景的图像
    img_width = (max_col * cell_width) + coord_width
    img_height = (max_row * cell_height) + coord_width
    img = Image.new("RGB", (img_width, img_height), color="white")

    # 创建绘图对象
    draw = ImageDraw.Draw(img)

    # 使用一个简单的字体
    font = ImageFont.load_default()

    # 绘制坐标
    for col in range(1, max_col + 1):
        x = coord_width + (col - 1) * cell_width
        draw.text((x + 5, 5), get_column_letter(col), fill="black", font=font)
    for row in range(1, max_row + 1):
        y = coord_width + (row - 1) * cell_height
        draw.text((5, y + 5), str(row), fill="black", font=font)

    # 处理合并的单元格
    merged_cells = ws.merged_cells.ranges

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)

            # 获取单元格位置
            x = coord_width + (col - 1) * cell_width
            y = coord_width + (row - 1) * cell_height

            # 检查是否是合并单元格的一部分
            merged = False
            for merged_range in merged_cells:
                if cell.coordinate in merged_range:
                    merged = True
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    x = coord_width + (min_col - 1) * cell_width
                    y = coord_width + (min_row - 1) * cell_height
                    width = (max_col - min_col + 1) * cell_width
                    height = (max_row - min_row + 1) * cell_height
                    break

            if not merged:
                width, height = cell_width, cell_height

            # 绘制单元格背景色
            if cell.fill.start_color.index != "00000000":
                fill_color = cell.fill.start_color.rgb
                if fill_color:
                    r, g, b = (
                        int(fill_color[2:4], 16),
                        int(fill_color[4:6], 16),
                        int(fill_color[6:8], 16),
                    )
                    draw.rectangle([x, y, x + width, y + height], fill=(r, g, b))

            # 绘制单元格边框
            draw.rectangle([x, y, x + width, y + height], outline="black")

            # 绘制单元格内容
            if cell.value and not (
                merged and cell.coordinate != f"{get_column_letter(min_col)}{min_row}"
            ):
                text = str(cell.value)
                left, top, right, bottom = font.getbbox(text)
                text_width = right - left
                text_height = bottom - top
                draw.text(
                    (x + (width - text_width) / 2, y + (height - text_height) / 2),
                    text,
                    fill="black",
                    font=font,
                )

    return img


# 使用示例
excel_file = "output_reversed_daily_month.xlsx"
sheet_name = "Sheet"  # 如果不指定，将使用活动工作表
img = excel_to_image(excel_file, sheet_name)
img.save("output3.png")
