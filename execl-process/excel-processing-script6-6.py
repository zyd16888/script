import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from collections import defaultdict
from calendar import month_abbr


# 获取颜色，根据数据值占最大值的百分比返回对应颜色代码
def get_color(value, max_value):
    if max_value == 0:
        return None
    percentage = (value / max_value) * 100
    if 90 <= percentage <= 100:
        return "8B4513"  # 棕色
    elif 80 <= percentage < 90:
        return "FF0000"  # 红色
    elif 70 <= percentage < 80:
        return "FFA500"  # 橙色
    elif 60 <= percentage < 70:
        return "FFFF00"  # 黄色
    elif 50 <= percentage < 60:
        return "008000"  # 绿色
    elif 40 <= percentage < 50:
        return "0000FF"  # 蓝色
    elif 0 < percentage < 40:
        return "00FFFF"
    else:
        return None


# 读取输入的 Excel 文件并初始化输出文件
def initialize_workbooks(input_file):
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    return ws_input, ws_output, wb_output


# 处理数据，按照小时和分钟整理数据并计算最大值
def process_data(ws_input):
    data = defaultdict(lambda: defaultdict(dict))
    dates = []
    max_value = 0

    for row in ws_input.iter_rows(min_row=2, values_only=True):
        date_str, *values = row
        date = parse_date(date_str)
        if not date:
            continue
        dates.append(date)

        for i, value in enumerate(values):
            if value is not None and isinstance(value, (int, float)):
                max_value = max(max_value, value)  # 更新最大值
            time = timedelta(minutes=5 * i)
            hour = (date + time).strftime("%H:00")
            minute = (date + time).strftime("%H:%M")
            data[hour][minute][date] = value

    dates.sort()  # 按日期排序
    return data, dates, max_value


# 将数据和格式写入输出的 Excel 文件
def write_data_to_excel(ws_output, data, dates, max_value):
    # 写入月份和日期标题
    write_headers(ws_output, dates)

    current_row = 3
    for hour in sorted(data.keys()):
        start_row = current_row
        for minute in sorted(data[hour].keys()):
            write_time_data(
                ws_output, hour, minute, data, dates, current_row, max_value
            )
            current_row += 1
        merge_hour_cells(ws_output, start_row, current_row)


chinese_months = [
    "一月",
    "二月",
    "三月",
    "四月",
    "五月",
    "六月",
    "七月",
    "八月",
    "九月",
    "十月",
    "十一月",
    "十二月",
]


# 写入表头，包括中文月份和日期
def write_headers(ws_output, dates):
    ws_output.insert_rows(1)
    col = 3
    current_month = None
    start_col = col

    # 写入中文月份合并单元格
    for date in dates:
        if current_month is None:
            current_month = date.month
        elif date.month != current_month:
            ws_output.merge_cells(
                start_row=1, start_column=start_col, end_row=1, end_column=col - 1
            )
            ws_output.cell(
                row=1, column=start_col, value=chinese_months[current_month - 1]
            )  # 使用中文月份
            current_month = date.month
            start_col = col
        col += 1

    # 合并最后一个月份
    ws_output.merge_cells(
        start_row=1, start_column=start_col, end_row=1, end_column=col - 1
    )
    ws_output.cell(
        row=1, column=start_col, value=chinese_months[current_month - 1]
    )  # 使用中文月份

    # 写入日期
    for col, date in enumerate(dates, start=3):
        ws_output.cell(row=2, column=col, value=date.strftime("%m/%d"))

    # 写入时间表头
    ws_output.cell(row=2, column=1, value="Hour")
    ws_output.cell(row=2, column=2, value="Minute")


# 将小时、分钟、数据和颜色写入 Excel 单元格
def write_time_data(ws_output, hour, minute, data, dates, row, max_value):
    ws_output.cell(row=row, column=1, value=hour)
    ws_output.cell(row=row, column=2, value=minute)

    for col, date in enumerate(dates, start=3):
        value = data[hour][minute].get(date)
        if value is not None:
            cell = ws_output.cell(row=row, column=col, value=value)
            color = get_color(value, max_value)
            if color:
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )


# 合并相同小时的单元格
def merge_hour_cells(ws_output, start_row, current_row):
    ws_output.merge_cells(
        start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1
    )


# 解析日期字符串，如果无法解析则返回 None
def parse_date(date_str):
    if isinstance(date_str, datetime):
        return date_str
    try:
        return datetime.strptime(date_str, "%m/%d")
    except ValueError:
        print(f"Warning: Could not parse date {date_str}. Skipping.")
        return None


# 主函数，调用其他函数进行 Excel 处理
def process_excel(input_file, output_file):
    ws_input, ws_output, wb_output = initialize_workbooks(input_file)
    data, dates, max_value = process_data(ws_input)
    write_data_to_excel(ws_output, data, dates, max_value)
    wb_output.save(output_file)


# 调用示例
input_file = r"新街口站255负荷.xlsx"
output_file = "output_reversed_daily_month_percentage2.xlsx"
process_excel(input_file, output_file)
