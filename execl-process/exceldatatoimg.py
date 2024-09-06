import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import numpy as np
from datetime import datetime
import os


def excel_to_images(excel_file, output_folder):
    # 创建输出文件夹
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 读取 Excel 文件
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 获取日期和时间
    dates = [cell.value for cell in ws[2][2:]]
    times = []
    data = []

    for row in ws.iter_rows(min_row=3, min_col=2, values_only=True):
        if row[0]:  # 检查时间列是否有值
            times.append(row[0])
            data.append(row[1:])

    # 将数据转换为 pandas DataFrame
    df = pd.DataFrame(data, index=times, columns=dates)

    # 为每个小时创建热力图
    for hour in range(24):
        hour_data = df[df.index.str.startswith(f"{hour:02d}:")]

        if hour_data.empty:
            continue

        fig, ax = plt.subplots(figsize=(20, 10))
        sns.heatmap(hour_data, cmap="YlOrRd", ax=ax, cbar_kws={"label": "Value"})

        ax.set_title(f"Heatmap for Hour {hour:02d}:00")
        ax.set_xlabel("Date")
        ax.set_ylabel("Time")

        # 调整 x 轴标签
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()

        # 保存图片
        output_file = os.path.join(output_folder, f"heatmap_hour_{hour:02d}.png")
        plt.savefig(output_file, dpi=300, bbox_inches="tight")
        plt.close()

    print(f"Images saved in {output_folder}")


# 使用示例
excel_file = "output_reversed_daily.xlsx"
output_folder = "heatmap_images"
excel_to_images(excel_file, output_folder)
